# ============================================================
# prepare_dataset.py  (v2 - đã cập nhật cho 11 ground truth)
# Mục đích: Đọc NLP.FSB_Pool.xlsx → tạo train.json, val.json, test.json
# Chiến lược với 11 GT:
#   - Train: 9 mẫu có GT  (81.8%)
#   - Val:   2 mẫu có GT  (18.2%)
#   - Test:  9 mẫu RAW (không có GT) → lấy từ phần "Còn lại"
# Chạy: python prepare_dataset.py
# ============================================================

import json, random, os, re
import openpyxl

EXCEL_PATH = "../NLP.FSB_Pool.xlsx"
OUTPUT_DIR = "."
SEED       = 42
random.seed(SEED)

# ── HÀM TRÍCH SPECS ───────────────────────────────────────────
def extract_specs(mo_ta_tho: str) -> dict:
    specs = {
        "duong": "", "phuong": "", "quan": "",
        "gia_ty": 0.0, "dien_tich_m2": 0.0, "so_tang": 0,
        "mat_tien_m": 0.0, "rong_hem_m": 0.0, "phan_loai_hem": "",
    }
    if not mo_ta_tho:
        return specs
    line1 = mo_ta_tho.strip().split("\n")[0]

    # Tên đường (trước số đầu tiên)
    m = re.match(r'^([A-Za-zÀ-ỹ\s\.\d]+?)\s+[\d]', line1)
    if m:
        duong_raw = m.group(1).strip()
        # Loại bỏ số ở cuối nếu có
        duong_raw = re.sub(r'\s+\d+$', '', duong_raw).strip()
        specs["duong"] = duong_raw

    # Diện tích
    m = re.search(r'(?:^|\s)([\d]+(?:\.[\d]+)?)(?:/[\d\.]+)?\s+\d+\s+[\d\.]+\s+[\d\.]+\s+[\d\.]+\s+tỷ', line1)
    if m:
        try: specs["dien_tich_m2"] = float(m.group(1))
        except: pass

    # Số tầng
    m = re.search(r'[\d\.]+(?:/[\d\.]+)?\s+(\d+)\s+[\d\.]+\s+[\d\.]+\s+[\d\.]+\s+tỷ', line1)
    if m:
        try: specs["so_tang"] = int(m.group(1))
        except: pass

    # Mặt tiền
    m = re.search(r'[\d\.]+(?:/[\d\.]+)?\s+\d+\s+([\d\.]+)\s+[\d\.]+\s+[\d\.]+\s+tỷ', line1)
    if m:
        try: specs["mat_tien_m"] = float(m.group(1))
        except: pass

    # Giá
    m = re.search(r'([\d]+(?:[,.][\d]+)?)\s*tỷ', line1)
    if m:
        try: specs["gia_ty"] = float(m.group(1).replace(',', '.'))
        except: pass

    # Phường & Quận
    m = re.search(r'Phường\s+([\w\s]+?)\s+Quận\s+(\w+)', mo_ta_tho)
    if m:
        specs["phuong"] = "Phường " + m.group(1).strip()
        specs["quan"]   = "Quận "   + m.group(2).strip()
    else:
        m = re.search(r'Quận\s+(\w+)', mo_ta_tho)
        if m: specs["quan"] = "Quận " + m.group(1).strip()

    # Phân loại hẻm
    full = mo_ta_tho.lower()
    if re.search(r'hẻm.*?xe tải|xe tải.*?hẻm', full):
        specs["phan_loai_hem"] = "Hẻm xe tải"
    elif re.search(r'hẻm.*?ô tô|ô tô.*?hẻm|hẻm.*?xe hơi|xe hơi.*?hẻm', full):
        specs["phan_loai_hem"] = "Hẻm xe hơi"
    elif 'hẻm' in full:
        specs["phan_loai_hem"] = "Hẻm"

    # Rộng hẻm
    m = re.search(r'hẻm\s+(?:rộng\s+)?([\d\.]+)\s*m', full)
    if m:
        try: specs["rong_hem_m"] = float(m.group(1))
        except: pass

    return specs

def clean(text):
    if not text: return ""
    return str(text).replace('\xa0', ' ').strip()

# ── ĐỌC EXCEL ─────────────────────────────────────────────────
print("Đoc file:", EXCEL_PATH)
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb['Pool']
print(f"Tong rows: {ws.max_row - 1}")

gt_rows  = []   # Có mô tả chuẩn → train/val
raw_rows = []   # Chỉ có mô tả thô → test

for row in range(2, ws.max_row + 1):
    mo_ta_tho     = clean(ws.cell(row, 1).value)
    system_id     = clean(ws.cell(row, 2).value)
    tieu_de_chuan = clean(ws.cell(row, 3).value)
    mo_ta_chuan   = clean(ws.cell(row, 4).value)

    if not mo_ta_tho or not system_id:
        continue

    record = {
        "id":                system_id,
        "raw_input_cleaned": mo_ta_tho,
        "clean_title":       tieu_de_chuan,
        "clean_description": mo_ta_chuan,
        "extracted_specs":   extract_specs(mo_ta_tho),
    }

    if mo_ta_chuan:
        gt_rows.append(record)
    else:
        raw_rows.append(record)

print(f"Ground truth rows: {len(gt_rows)}")
print(f"Raw-only rows:     {len(raw_rows)}")

# ── CHIA TRAIN / VAL / TEST ───────────────────────────────────
# Với 11 GT: Train=9, Val=2
# Test: 9 mẫu lấy từ raw_rows (không có nhãn)
random.shuffle(gt_rows)
random.shuffle(raw_rows)

train_data = gt_rows[:9]
val_data   = gt_rows[9:]       # 2 mẫu còn lại
test_data  = raw_rows[:9]      # 9 mẫu RAW (không có GT label)

print(f"\nPhan chia:")
print(f"  Train: {len(train_data)} mau (co ground truth)")
print(f"  Val:   {len(val_data)} mau (co ground truth)")
print(f"  Test:  {len(test_data)} mau (KHONG co GT - danh gia dinh tinh)")

# ── LƯU JSON ──────────────────────────────────────────────────
os.makedirs(OUTPUT_DIR, exist_ok=True)

for data, fname in [(train_data, "train.json"),
                    (val_data,   "val.json"),
                    (test_data,  "test.json")]:
    path = os.path.join(OUTPUT_DIR, fname)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"Saved: {path} ({len(data)} records)")

# ── IN MẪU KIỂM TRA ───────────────────────────────────────────
print("\n--- MAU TRAIN[0] ---")
s = train_data[0]
print(f"ID:    {s['id']}")
print(f"Input: {s['raw_input_cleaned'][:100]}...")
print(f"GT:    {s['clean_description'][:100]}...")
print(f"Specs: {s['extracted_specs']}")

print("\n--- MAU TEST[0] (khong co GT) ---")
s = test_data[0]
print(f"ID:    {s['id']}")
print(f"Input: {s['raw_input_cleaned'][:100]}...")

print("\nDone! Upload train.json, val.json, test.json len Colab.")
print("Luu y: Test set khong co GT nen ROUGE se khong tinh duoc chinh xac.")
print("Trang (Leader) se danh gia test set dinh tinh (Expert Preference).")
